from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.assistant.orchestrator import AssistantOrchestrator
from app.main import app
from app.tools.files.file_creator import FileCreatorTool
from app.tools.files.path_safety import safe_join, sanitize_filename


client = TestClient(app)


def test_create_excel_file_in_export_folder(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    result = FileCreatorTool().create_excel_file(
        "Ausgaben",
        [{"name": "Ausgaben", "headers": ["Datum", "Betrag"], "rows": [["2026-06-05", 12.5]]}],
        filename="ausgaben.xlsx",
    )

    assert result["created"] is True
    assert result["file_type"] == "xlsx"
    assert Path(result["path"]).is_file()
    assert str(tmp_path / "exports") in result["path"]


def test_excel_headers_are_created(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    result = FileCreatorTool().create_excel_file(
        "Plan",
        [{"name": "Plan", "headers": ["Maschine", "Status"], "rows": [["Presse", "OK"]]}],
        filename="plan.xlsx",
    )

    workbook = load_workbook(result["path"])
    sheet = workbook["Plan"]
    assert sheet["A1"].value == "Maschine"
    assert sheet["B1"].value == "Status"


def test_filename_is_sanitized() -> None:
    assert sanitize_filename("a<>b:c?.xlsx") == "a_b_c_.xlsx"


def test_path_traversal_is_rejected(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    try:
        safe_join(tmp_path / "exports", "../bad.xlsx")
    except ValueError as exc:
        assert "ungueltig" in str(exc)
    else:
        raise AssertionError("path traversal must fail")


def test_duplicate_file_creates_suffix(monkeypatch, tmp_path) -> None:
    export_dir = tmp_path / "exports"
    monkeypatch.setenv("EXPORT_DIR", str(export_dir))
    export_dir.mkdir(parents=True)
    (export_dir / "ausgaben.xlsx").write_text("exists", encoding="utf-8")

    result = FileCreatorTool().create_excel_file(
        "Ausgaben",
        [{"name": "Ausgaben", "headers": ["Datum"], "rows": []}],
        filename="ausgaben.xlsx",
    )

    assert result["filename"] == "ausgaben_001.xlsx"


def test_create_csv_works(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    result = FileCreatorTool().create_csv_file(["A", "B"], [[1, 2]], filename="daten.csv")

    assert result["created"] is True
    assert Path(result["path"]).read_text(encoding="utf-8").splitlines()[0] == "A,B"


def test_create_markdown_works(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    result = FileCreatorTool().create_markdown_file("Titel", "Inhalt", filename="notiz.md")

    assert result["created"] is True
    assert "# Titel" in Path(result["path"]).read_text(encoding="utf-8")


def test_create_excel_endpoint_returns_200(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    response = client.post(
        "/assistant/files/create/excel",
        json={
            "title": "Ausgaben",
            "filename": "ausgaben.xlsx",
            "sheets": [{"name": "Ausgaben", "headers": ["Datum"], "rows": []}],
        },
    )

    assert response.status_code == 200
    assert response.json()["created"] is True


def test_assistant_excel_expenses_intent_creates_file(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    result = AssistantOrchestrator().handle_message("erstelle eine Excel fuer Ausgaben")

    assert result["tool"] == "file_create_excel"
    assert result["result"]["created"] is True
    assert result["result"]["path"] in result["answer"]


def test_assistant_spoken_excel_expenses_intent_creates_workbook(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    result = AssistantOrchestrator().handle_message("Jarvis, erstelle mir eine Excel für meine Ausgaben.")

    assert result["tool"] == "file_create_excel"
    assert result["created"] is True
    assert result["file_type"] == "xlsx"
    assert result["filename"] == "ausgaben.xlsx"
    assert result["path"] in result["answer"]
    assert "Ich habe die Excel-Datei erstellt:" in result["answer"]
    assert "Ich kann keine Excel-Dateien" not in result["answer"]
    assert "keine Spreadsheet-Tools" not in result["answer"]

    workbook = load_workbook(result["path"])
    sheet = workbook["Ausgaben"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 7)]
    assert headers == ["Datum", "Kategorie", "Beschreibung", "Betrag", "Zahlungsart", "Notiz"]


def test_assistant_spoken_excel_expenses_duplicate_uses_suffix(monkeypatch, tmp_path) -> None:
    export_dir = tmp_path / "exports"
    monkeypatch.setenv("EXPORT_DIR", str(export_dir))
    export_dir.mkdir(parents=True)
    (export_dir / "ausgaben.xlsx").write_text("exists", encoding="utf-8")

    result = AssistantOrchestrator().handle_message("Jarvis, erstelle mir eine Excel für meine Ausgaben.")

    assert result["filename"] == "ausgaben_001.xlsx"
    assert Path(result["path"]).name == "ausgaben_001.xlsx"


def test_file_exports_endpoint_returns_200(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))

    response = client.get("/assistant/files/exports")

    assert response.status_code == 200
    assert "files" in response.json()
