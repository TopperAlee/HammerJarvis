import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.tools.files.path_safety import get_export_dir, safe_join, sanitize_filename


class FileCreatorTool:
    def create_excel_file(
        self,
        title: str,
        sheets: list[dict[str, Any]],
        filename: str | None = None,
    ) -> dict[str, Any]:
        output = _output_path(filename, title, ".xlsx")
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for sheet_data in sheets or [_empty_sheet(title)]:
            sheet_name = _safe_sheet_name(str(sheet_data.get("name") or title or "Sheet1"))
            sheet = workbook.create_sheet(sheet_name)
            headers = [str(value) for value in sheet_data.get("headers", [])]
            rows = sheet_data.get("rows", [])
            if headers:
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                sheet.freeze_panes = "A2"
            for row in rows:
                sheet.append(list(row))
            _format_excel_sheet(sheet)
        workbook.save(output)
        return _created("xlsx", output, f"Excel-Datei wurde erstellt: {output.name}")

    def create_csv_file(
        self,
        headers: list[str],
        rows: list[list[Any]],
        filename: str | None = None,
    ) -> dict[str, Any]:
        output = _output_path(filename, "export", ".csv")
        with output.open("w", encoding="utf-8", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            writer.writerows(rows)
        return _created("csv", output, f"CSV-Datei wurde erstellt: {output.name}")

    def create_markdown_file(
        self,
        title: str,
        content: str,
        filename: str | None = None,
    ) -> dict[str, Any]:
        output = _output_path(filename, title, ".md")
        output.write_text(f"# {title}\n\n{content}\n", encoding="utf-8")
        return _created("md", output, f"Markdown-Datei wurde erstellt: {output.name}")

    def create_json_file(self, data: dict[str, Any], filename: str | None = None) -> dict[str, Any]:
        output = _output_path(filename, "daten", ".json")
        with output.open("w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)
            file.write("\n")
        return _created("json", output, f"JSON-Datei wurde erstellt: {output.name}")

    def list_exports(self) -> dict[str, Any]:
        export_dir = get_export_dir()
        export_dir.mkdir(parents=True, exist_ok=True)
        files = [
            {
                "filename": path.name,
                "path": str(path),
                "size_bytes": path.stat().st_size,
            }
            for path in sorted(export_dir.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True)
            if path.is_file() and path.name != ".gitkeep"
        ]
        return {"export_dir": str(export_dir), "files": files}


def _output_path(filename: str | None, title: str, suffix: str) -> Path:
    base_name = filename or f"{sanitize_filename(title)}{suffix}"
    if not base_name.lower().endswith(suffix):
        base_name += suffix
    return safe_join(get_export_dir(), base_name)


def _empty_sheet(title: str) -> dict[str, Any]:
    return {"name": title or "Sheet1", "headers": [], "rows": []}


def _safe_sheet_name(name: str) -> str:
    cleaned = sanitize_filename(name).replace("_", " ")[:31].strip()
    return cleaned or "Sheet1"


def _format_excel_sheet(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 60)
        for cell in column_cells:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"


def _created(file_type: str, path: Path, message: str) -> dict[str, Any]:
    return {
        "created": True,
        "file_type": file_type,
        "filename": path.name,
        "path": str(path),
        "message": message,
    }
